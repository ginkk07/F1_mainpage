#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CSV → Excel 轉換工具
將 seat-data.csv 轉換為 seat-data.xlsx（Excel 格式）
完全避免編碼問題

安裝依賴：
  pip install openpyxl
  或
  pip install pandas openpyxl
"""

import csv
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("❌ 需要安裝 openpyxl")
    print("   運行：pip install openpyxl")
    sys.exit(1)

def csv_to_excel():
    script_dir = Path(__file__).parent
    project_root = script_dir.parent
    csv_path = project_root / "seat-data.csv"
    excel_path = project_root / "seat-data.xlsx"
    
    if not csv_path.exists():
        print(f"❌ 找不到 CSV 檔案：{csv_path}")
        return
    
    # 讀取 CSV（支持 BOM）
    rows = []
    try:
        with open(csv_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            for row in reader:
                rows.append(row)
    except Exception as e:
        print(f"❌ 讀取 CSV 失敗：{e}")
        return
    
    if not rows:
        print("❌ CSV 檔案為空")
        return
    
    # 創建 Excel 工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "座位資料"
    
    # 寫入數據
    for row_idx, row_data in enumerate(rows, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 格式化表頭
            if row_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 調整列寬
    column_widths = {
        'A': 22,  # ID
        'B': 30,  # 標題
        'C': 12,  # 分類
        'D': 15,  # 區域
        'E': 25,  # SVG Map ID
        'F': 35,  # 圖片路徑
        'G': 60,  # 描述
        'H': 15,  # 備註
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # 自動換行
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # 凍結表頭
    ws.freeze_panes = "A2"
    
    # 保存
    try:
        wb.save(excel_path)
        print(f"✅ 轉換成功！")
        print(f"📁 輸出文件：{excel_path}")
        print(f"📊 共轉換 {len(rows) - 1} 個座位（+ 1 行表頭）")
        print(f"\n💡 使用建議：")
        print(f"   - 用 Excel 直接打開 {excel_path.name}")
        print(f"   - 編輯完後保存")
        print(f"   - 再用 CSV 轉換工具轉回 JavaScript 代碼")
    except Exception as e:
        print(f"❌ 保存失敗：{e}")

if __name__ == "__main__":
    csv_to_excel()
